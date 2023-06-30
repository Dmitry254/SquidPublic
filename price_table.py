import openpyxl


def fill_table_buses(buses):
    wb = open_table("")
    sheet_name = ''
    recreate_sheet(wb, sheet_name)
    sheet = wb[sheet_name]
    row = sheet.max_row
    if row == 1:
        row -= 2
    try:
        for level in range(0, 10):
            list_name = 'bus' + str(level)
            if list_name in buses.keys():
                row += 2
                sheet.cell(row=row, column=1).value = f"Level {level}"
                prices_list = buses[list_name]
                row += 1
                for price_number in range(len(prices_list)):
                    sheet.cell(row=row, column=price_number+1).value = prices_list[price_number]
                row += 1
                prices_list = sorted(buses[list_name])
                for price_number in range(len(prices_list)):
                    sheet.cell(row=row, column=price_number+1).value = prices_list[price_number]
            else:
                continue
    finally:
        close_table(wb, "history")


def fill_table_players(players):
    wb = open_table("history")
    sheet_name = 'prices'
    sheet = wb[sheet_name]
    row = sheet.max_row
    if row == 1:
        row -= 2
    try:
        for level in range(1, 6):
            for energy in range(0, 45):
                list_name = str(level) + 'player' + str(energy)
                if list_name in players.keys():
                    row += 2
                    sheet.cell(row=row, column=1).value = f"Energy {energy}00-{energy+1}00, level {level}"
                    prices_list = players[list_name]
                    row += 1
                    for price_number in range(len(prices_list)):
                        sheet.cell(row=row, column=price_number+1).value = prices_list[price_number]
                    row += 1
                    prices_list = sorted(players[list_name])
                    for price_number in range(len(prices_list)):
                        sheet.cell(row=row, column=price_number+1).value = prices_list[price_number]
    finally:
        close_table(wb, "history")


def fill_table_robi(robies):
    wb = open_table("history")
    sheet_name = 'prices'
    sheet = wb[sheet_name]
    row = sheet.max_row
    if row == 1:
        row -= 2
    try:
        for level in range(0, 10):
            list_name = 'robi' + str(level)
            if list_name in robies.keys():
                row += 2
                sheet.cell(row=row, column=1).value = f"Robi {level}"
                robi_list = robies[list_name]
                row += 1
                for robi_number in range(len(robi_list)):
                    sheet.cell(row=row, column=robi_number+1).value = robi_list[robi_number]
            else:
                continue
    finally:
        close_table(wb, "history")


def open_table(table_name):
    wb = openpyxl.load_workbook(f"{table_name}.xlsx")
    return wb


def close_table(wb, table_name):
    wb.save(f'{table_name}.xlsx')


def recreate_sheet(wb, sheet_name):
    sheet = wb[sheet_name]
    wb.remove(sheet)
    wb.create_sheet(sheet_name)


def fill_energy_value():
    wb = open_table()
    sheet = wb['price']
    for level in range(0, 40):
        sheet.cell(row=3+level, column=5).value = f"Energy {level}00-{level + 1}00"
    close_table(wb)
